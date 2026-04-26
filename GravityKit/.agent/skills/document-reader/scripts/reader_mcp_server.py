#!/usr/bin/env python3
"""
reader_mcp_server.py — MCP server exposing document reading capabilities.

Tools:
    read_document(filepath: str) -> str
"""
import sys
from pathlib import Path

try:
    from mcp.server.fastmcp import FastMCP
except ImportError:
    FastMCP = None


def read_pdf(path: Path) -> str:
    try:
        import pypdf
    except ImportError:
        return "Error: pypdf is not installed. Please run: pip install pypdf"
        
    with open(path, "rb") as f:
        reader = pypdf.PdfReader(f)
        text = "\n\n".join(page.extract_text() for page in reader.pages if page.extract_text())
    return text


def read_docx(path: Path) -> str:
    try:
        import docx
    except ImportError:
        return "Error: python-docx is not installed. Please run: pip install python-docx"
        
    doc = docx.Document(path)
    return "\n".join(para.text for para in doc.paragraphs)


def read_xlsx(path: Path) -> str:
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl is not installed. Please run: pip install openpyxl"
        
    wb = openpyxl.load_workbook(path, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"## Sheet: {sheet}")
        for row in ws.iter_rows(values_only=True):
            row_str = " | ".join(str(cell).replace("\n", " ") if cell is not None else "" for cell in row)
            lines.append(f"| {row_str} |")
        lines.append("")
    return "\n".join(lines)


def read_document(filepath: str) -> str:
    """Reads the text content of a PDF, DOCX, or XLSX file and returns it as Markdown format. Use this whenever the user asks to process or read a document."""
    path = Path(filepath).resolve()
    if not path.exists():
        return f"Error: File not found at {filepath}"
    
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            return read_pdf(path)
        elif ext == ".docx":
            return read_docx(path)
        elif ext == ".xlsx":
            return read_xlsx(path)
        else:
            return f"Error: Unsupported file extension {ext}. Supported: .pdf, .docx, .xlsx"
    except Exception as e:
        return f"Error reading document: {str(e)}"


def main():
    if FastMCP is None:
        print("Missing mcp package. pip install mcp", file=sys.stderr)
        sys.exit(1)
        
    mcp = FastMCP("document-reader")
    mcp.tool()(read_document)
    mcp.run(transport="stdio")


if __name__ == "__main__":
    main()
